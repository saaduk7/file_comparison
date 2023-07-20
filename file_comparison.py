import json
from helpers.classes import *
from helpers.image_comparison import *
from os import listdir
import json
from os.path import isfile, join, splitext
import configparser
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, Color
from helpers.adv_image_comparison import compare_image_files


def create_less_detailed_differnces(diff_list):
    less_detailed_diff = Differences()
    for diff in diff_list:
        if "Curve Color" in diff:
            less_detailed_diff.curve_color = True
        elif "Different Number of Tables" in diff:
            less_detailed_diff.layout = True
        elif "X_Scaling" in diff:
            less_detailed_diff.x_axis_range = True
        elif "Y_Scaling" in diff:
            less_detailed_diff.y_axis_range = True
        elif "Legend Value" in diff:
            less_detailed_diff.legend_values = True
        elif "Missing Curve" in diff:
            less_detailed_diff.missing_curves = True
        elif "Image shift" in diff:
            less_detailed_diff.shifted_pixels = True    
        
            
        
    return [k for k, v in vars(less_detailed_diff).items() if v]


"""
main file of file comparison tool. Uses the config.txt to compare the two reports given. Creates an excel file containing differences.
"""

config = configparser.RawConfigParser()
config.read('config.txt')
config_dict = dict(config.items('FOLDER_PATHS'))
detailed_report = config_dict["detailed_report"]
output_file = config_dict["output_path"]
onlyfiles = sorted([f for f in listdir(config_dict["folder_1"]) if f.endswith(".json")])
onlyfiles_2 = sorted([f for f in listdir(config_dict["folder_2"]) if f.endswith(".json")])


wb = Workbook()
# grab the active worksheet
ws = wb.active
a1 = ws['A1']
a1.font = Font(bold=True)

for file in onlyfiles:
    if file not in onlyfiles_2:
        ws['A1'] = f"FOLDER MISMATCH number of files: folder1: {len(onlyfiles)}, folder2: {len(onlyfiles_2)}"
        print(f"FOLDER MISMATCH number of files: folder1: {len(onlyfiles)}, folder2: {len(onlyfiles_2)}")
        # close the file
        wb.save(output_file)
        exit()
onlyimages = sorted([f for f in listdir(config_dict["folder_1"]) if f.endswith(".EMF")])
print(config_dict)


ws['A1'] = f"Comparing {config_dict['folder_1']} with {config_dict['folder_2']}"
ws["A2"] = "Filename"
ws["B2"] = "Differences"

for file in onlyfiles:
    print(f"Comparing file {file}")
    with open(join(config_dict["folder_1"], file)) as f:
        json_report_1 = JsonReport(config_dict["folder_1"]+file, json.load(f))
    with open(join(config_dict["folder_2"], file)) as f:
        json_report_2 = JsonReport(config_dict["folder_2"]+file, json.load(f))
    diff_list = json_report_1 == json_report_2
    
    diff = compare_images(join(config_dict["folder_1"], splitext(file)[0] + ".EMF"), join(config_dict["folder_2"], splitext(file)[0] + ".EMF"))
    if diff:    
        diff_list.append(diff)
    
    diff = compare_image_files(join(config_dict["folder_1"], splitext(file)[0] + ".EMF"), join(config_dict["folder_2"], splitext(file)[0] + ".EMF"))
    if diff:    
        diff_list.extend(diff)
                             
    if detailed_report.lower()=="false":
        diff_list = create_less_detailed_differnces(diff_list)
    # create the csv writer
    if diff_list:
        ws.append([splitext(file)[0], "Following Differences Found"])
        for diff in diff_list:   
            ws.append(["", diff])
    else:
        ws.append([file, "No Differences Found"])
# close the file
wb.save(output_file)
